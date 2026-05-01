#!/usr/bin/env python3
"""Shared helpers for the weekly pipeline scripts."""

from __future__ import annotations

import json
import os
import re
import subprocess
import socket
import ssl
import time
from collections.abc import Iterable
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any

PROJECT_ROOT = Path(__file__).resolve().parents[1]
PIPELINE_ROOT = Path(__file__).resolve().parent
CONFIG_PATH = PIPELINE_ROOT / "pipeline-config.json"
STATE_PATH = PIPELINE_ROOT / "pipeline-state.json"
PIPELINE_OVERRIDES_PATH = PIPELINE_ROOT / "pipeline-overrides.json"
PROCESSED_DRIVE_WORKBOOKS_PATH = PIPELINE_ROOT / "processed-drive-workbooks.json"
DOWNLOAD_ROOT = PIPELINE_ROOT / "output" / "downloaded-workbooks"
EXTRACTED_ROOT = PIPELINE_ROOT / "output" / "extracted-csv"
ARCHIVE_ROOT = PROJECT_ROOT / "dataGoogleDrive"
IMPORT_SCRIPT = PROJECT_ROOT / "pipeline" / "import-google-drive-folder.py"
ACTIVE_DRIVE_YEAR_FOLDER = "2026"

GOOGLE_FOLDER_MIME = "application/vnd.google-apps.folder"
GOOGLE_SHEETS_MIME = "application/vnd.google-apps.spreadsheet"
GOOGLE_SHORTCUT_MIME = "application/vnd.google-apps.shortcut"
XLSX_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
DRIVE_READONLY_SCOPE = "https://www.googleapis.com/auth/drive.readonly"
SUPPORTED_FILE_MIME_TYPES = {GOOGLE_SHEETS_MIME, XLSX_MIME}
WINDOWS_RESERVED_CHARS = re.compile(r'[<>:"/\\|?*]')
DEFAULT_COMMIT_PATHS = [
    PROJECT_ROOT / "data" / "events",
    PROJECT_ROOT / "data" / "elo-data",
    PROJECT_ROOT / "data" / "events.json",
    PROJECT_ROOT / "data" / "results.json",
    PROJECT_ROOT / "data" / "aliases.json",
    PROJECT_ROOT / "data" / "matchups",
    PIPELINE_OVERRIDES_PATH,
    PROCESSED_DRIVE_WORKBOOKS_PATH,
]

DRIVE_REQUEST_METRICS: dict[str, int] = {
    "total_requests": 0,
    "files_list_requests": 0,
    "files_get_requests": 0,
    "media_download_requests": 0,
}


@dataclass(frozen=True)
class PipelineSettings:
    credentials_path: Path
    drive_folder_id: str
    download_root: Path
    extracted_root: Path
    archive_root: Path
    import_script: Path
    remote: str
    data_branch: str
    main_branch: str
    commit_message_template: str


@dataclass(frozen=True)
class DriveFile:
    file_id: str
    name: str
    mime_type: str
    modified_time: datetime
    folder_segments: tuple[str, ...]

    @property
    def export_name(self) -> str:
        if self.mime_type == GOOGLE_SHEETS_MIME and not self.name.lower().endswith(".xlsx"):
            return f"{self.name}.xlsx"
        return self.name


def log(message: str) -> None:
    print(message)


def reset_drive_request_metrics() -> None:
    for key in DRIVE_REQUEST_METRICS:
        DRIVE_REQUEST_METRICS[key] = 0


def record_drive_request(metric_name: str) -> None:
    DRIVE_REQUEST_METRICS["total_requests"] += 1
    if metric_name in DRIVE_REQUEST_METRICS:
        DRIVE_REQUEST_METRICS[metric_name] += 1


def get_drive_request_metrics() -> dict[str, int]:
    return dict(DRIVE_REQUEST_METRICS)


def load_json_file(path: Path, *, default: dict[str, Any] | None = None) -> dict[str, Any]:
    if not path.exists():
        return {} if default is None else dict(default)
    try:
        return json.loads(path.read_text(encoding="utf-8"))
    except json.JSONDecodeError as exc:
        raise SystemExit(f"JSON file is not valid: {path}\n{exc}") from exc


def resolve_config_path(value: object, fallback: Path) -> Path:
    if value in (None, ""):
        return fallback.resolve()

    candidate = Path(str(value)).expanduser()
    if not candidate.is_absolute():
        candidate = CONFIG_PATH.parent / candidate
    return candidate.resolve()


def load_settings() -> PipelineSettings:
    if not CONFIG_PATH.exists():
        raise SystemExit(
            "Missing pipeline config.\n"
            f"Create '{CONFIG_PATH.name}' next to this script, or run through the GitHub Actions wrapper that writes it temporarily."
        )

    config = load_json_file(CONFIG_PATH)
    credentials_value = config.get("credentials_path") or os.environ.get("GOOGLE_SERVICE_ACCOUNT_FILE")
    drive_folder_value = config.get("drive_folder_id") or os.environ.get("GOOGLE_DRIVE_FOLDER_ID")
    if not credentials_value:
        raise SystemExit(
            "Missing credentials_path.\n"
            "Fill it in pipeline-config.json or set GOOGLE_SERVICE_ACCOUNT_FILE."
        )
    if not drive_folder_value:
        raise SystemExit(
            "Missing drive_folder_id.\n"
            "Fill it in pipeline-config.json or set GOOGLE_DRIVE_FOLDER_ID."
        )

    return PipelineSettings(
        credentials_path=resolve_config_path(credentials_value, Path(str(credentials_value)).expanduser()),
        drive_folder_id=str(drive_folder_value).strip(),
        download_root=resolve_config_path(config.get("download_root"), DOWNLOAD_ROOT),
        extracted_root=resolve_config_path(config.get("extracted_root"), EXTRACTED_ROOT),
        archive_root=resolve_config_path(config.get("archive_root"), ARCHIVE_ROOT),
        import_script=resolve_config_path(config.get("import_script"), IMPORT_SCRIPT),
        remote=str(config.get("remote", "origin")),
        data_branch=str(config.get("data_branch") or config.get("development_branch") or "data-updates"),
        main_branch=str(config.get("main_branch", "main")),
        commit_message_template=str(
            config.get("commit_message_template", "chore(data): import {workbook_name}")
        ),
    )


def load_state() -> dict[str, Any]:
    return load_json_file(STATE_PATH, default={})


def save_state(payload: dict[str, Any]) -> None:
    STATE_PATH.parent.mkdir(parents=True, exist_ok=True)
    STATE_PATH.write_text(json.dumps(payload, indent=2) + "\n", encoding="utf-8")


def require_state_keys(*keys: str) -> dict[str, Any]:
    state = load_state()
    missing = [key for key in keys if key not in state or state.get(key) in (None, "")]
    if missing:
        missing_list = ", ".join(missing)
        raise SystemExit(
            f"State is missing: {missing_list}\n"
            "Run the earlier automated pipeline steps first."
        )
    return state


def sanitize_path_part(value: str) -> str:
    sanitized = WINDOWS_RESERVED_CHARS.sub("_", value).strip().rstrip(". ")
    return sanitized or "_"


def parse_drive_time(raw_value: str) -> datetime:
    return datetime.fromisoformat(raw_value.replace("Z", "+00:00"))


def normalize_cell_text(value: object) -> str:
    if value is None:
        return ""
    return str(value).strip()


def summarize_missing_positions(positions: Iterable[int], *, limit: int = 8) -> str:
    unique_positions = sorted({int(position) for position in positions})
    preview = ", ".join(str(position) for position in unique_positions[:limit])
    if len(unique_positions) > limit:
        preview = f"{preview}, ..."
    return preview


def inspect_workbook_readiness(workbook_path: Path) -> dict[str, Any]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise SystemExit(
            "Workbook readiness checks require openpyxl.\n"
            "Install it with:\n"
            "  python -m pip install openpyxl"
        ) from exc

    issues: list[str] = []
    input_missing_positions: list[int] = []
    matchup_has_data = False

    workbook = load_workbook(filename=str(workbook_path), read_only=True, data_only=True)
    try:
        if "Input" not in workbook.sheetnames:
            issues.append("Input sheet is missing.")
        else:
            input_sheet = workbook["Input"]
            for top32_position in range(1, 33):
                row_index = top32_position + 1 #header is row 1; players in column B (2) and decks in column E (5)
                name_value = normalize_cell_text(input_sheet.cell(row=row_index, column=2).value)
                deck_value = normalize_cell_text(input_sheet.cell(row=row_index, column=5).value)
                if not name_value or not deck_value:
                    input_missing_positions.append(top32_position)
            for row_index in range(34, input_sheet.max_row + 1 ):
                name_value = normalize_cell_text(input_sheet.cell(row=row_index, column=2).value)
                deck_value = normalize_cell_text(input_sheet.cell(row=row_index, column=5).value)
                if bool(name_value) != bool(deck_value):
                    issues.append(f"Input has mismatched Name/Deck rows at position {row_index}")
                    break

            if input_missing_positions:
                issues.append(
                    "Input Top 32 is incomplete for Name/Deck rows: "
                    f"{summarize_missing_positions(input_missing_positions)}"
                )
        if "Match Up Input" not in workbook.sheetnames:
            issues.append("Match Up Input sheet is missing.")
        else:
            matchup_sheet = workbook["Match Up Input"]
            for row in matchup_sheet.iter_rows(min_row=2, values_only=True):
                if any(normalize_cell_text(value) for value in row):
                    matchup_has_data = True
                    break

            if not matchup_has_data:
                issues.append("Match Up Input sheet has no data rows yet.")
    finally:
        workbook.close()

    return {
        "is_ready": len(issues) == 0,
        "issues": issues,
        "input_top32_complete": len(input_missing_positions) == 0,
        "input_missing_positions": input_missing_positions,
        "matchup_has_data": matchup_has_data,
    }


def relative_posix(path: Path, base: Path) -> str:
    return path.relative_to(base).as_posix()


def relative_archive_path(drive_file: DriveFile) -> Path:
    safe_segments = [sanitize_path_part(segment) for segment in drive_file.folder_segments]
    safe_name = sanitize_path_part(drive_file.export_name)
    return Path(*safe_segments, safe_name) if safe_segments else Path(safe_name)


def legacy_archive_relative_path(drive_file: DriveFile) -> Path:
    return Path(
        *[
            part.replace("'", "_")
            for part in relative_archive_path(drive_file).parts
        ]
    )


def archive_relative_path_candidates(drive_file: DriveFile) -> list[Path]:
    current = relative_archive_path(drive_file)
    legacy = legacy_archive_relative_path(drive_file)
    candidates = [current]
    if legacy != current:
        candidates.append(legacy)
    return candidates


def resolve_archive_relative_path(
    drive_file: DriveFile, archive_root: Path, archive_paths: set[str]
) -> Path:
    for candidate in archive_relative_path_candidates(drive_file):
        if candidate.as_posix() in archive_paths:
            return candidate

    legacy = legacy_archive_relative_path(drive_file)
    if (archive_root / legacy.parent).exists():
        return legacy

    return relative_archive_path(drive_file)


def build_drive_service(credentials_path: Path):
    try:
        from google.oauth2.service_account import Credentials
        from googleapiclient.discovery import build
    except ImportError as exc:
        raise SystemExit(
            "Google Drive libraries are missing.\n"
            "Install them with:\n"
            "  python -m pip install -r .\\pipeline\\requirements.txt"
        ) from exc

    if not credentials_path.exists():
        raise SystemExit(f"Credentials file was not found: {credentials_path}")

    credentials = Credentials.from_service_account_file(
        str(credentials_path),
        scopes=[DRIVE_READONLY_SCOPE],
    )
    return build("drive", "v3", credentials=credentials, cache_discovery=False)


def list_folder_children(service, folder_id: str) -> list[dict[str, Any]]:
    items: list[dict[str, Any]] = []
    page_token: str | None = None

    while True:
        response = (
            service.files()
            .list(
                q=f"'{folder_id}' in parents and trashed = false",
                fields=(
                    "nextPageToken,"
                    "files(id,name,mimeType,modifiedTime,shortcutDetails/targetId,shortcutDetails/targetMimeType)"
                ),
                orderBy="folder,name",
                pageSize=1000,
                pageToken=page_token,
                supportsAllDrives=True,
                includeItemsFromAllDrives=True,
            )
            .execute()
        )
        record_drive_request("files_list_requests")
        items.extend(response.get("files", []))
        page_token = response.get("nextPageToken")
        if not page_token:
            break

    return items


def fetch_drive_item(service, file_id: str) -> dict[str, Any]:
    response = (
        service.files()
        .get(
            fileId=file_id,
            fields="id,name,mimeType,modifiedTime",
            supportsAllDrives=True,
        )
        .execute()
    )
    record_drive_request("files_get_requests")
    return response


def resolve_shortcut(service, item: dict[str, Any]) -> dict[str, Any] | None:
    shortcut = item.get("shortcutDetails") or {}
    target_id = shortcut.get("targetId")
    target_mime_type = shortcut.get("targetMimeType")
    if not target_id or target_mime_type not in ({GOOGLE_FOLDER_MIME} | SUPPORTED_FILE_MIME_TYPES):
        return None

    resolved = fetch_drive_item(service, str(target_id))
    return {
        "id": resolved["id"],
        "name": resolved["name"],
        "mimeType": resolved["mimeType"],
        "modifiedTime": resolved["modifiedTime"],
    }


def find_latest_drive_file(service, root_folder_id: str) -> DriveFile:
    latest: DriveFile | None = None
    queue: list[tuple[str, tuple[str, ...]]] = [(root_folder_id, ())]
    seen_folders: set[str] = set()

    while queue:
        folder_id, folder_segments = queue.pop()
        if folder_id in seen_folders:
            continue
        seen_folders.add(folder_id)

        for item in list_folder_children(service, folder_id):
            candidate = item
            if item.get("mimeType") == GOOGLE_SHORTCUT_MIME:
                resolved = resolve_shortcut(service, item)
                if resolved is None:
                    continue
                candidate = resolved

            mime_type = str(candidate.get("mimeType", ""))
            if mime_type == GOOGLE_FOLDER_MIME:
                queue.append((str(candidate["id"]), folder_segments + (str(candidate["name"]),)))
                continue
            if mime_type not in SUPPORTED_FILE_MIME_TYPES:
                continue

            drive_file = DriveFile(
                file_id=str(candidate["id"]),
                name=str(candidate["name"]),
                mime_type=mime_type,
                modified_time=parse_drive_time(str(candidate["modifiedTime"])),
                folder_segments=folder_segments,
            )
            if latest is None or drive_file.modified_time > latest.modified_time:
                latest = drive_file

    if latest is None:
        raise SystemExit("No Google Sheets or .xlsx files were found under the configured Drive folder.")

    return latest


def list_drive_files(service, root_folder_id: str) -> list[DriveFile]:
    drive_files: list[DriveFile] = []
    queue: list[tuple[str, tuple[str, ...]]] = [(root_folder_id, ())]
    seen_folders: set[str] = set()

    while queue:
        folder_id, folder_segments = queue.pop()
        if folder_id in seen_folders:
            continue
        seen_folders.add(folder_id)

        for item in list_folder_children(service, folder_id):
            candidate = item
            if item.get("mimeType") == GOOGLE_SHORTCUT_MIME:
                resolved = resolve_shortcut(service, item)
                if resolved is None:
                    continue
                candidate = resolved

            mime_type = str(candidate.get("mimeType", ""))
            if mime_type == GOOGLE_FOLDER_MIME:
                queue.append((str(candidate["id"]), folder_segments + (str(candidate["name"]),)))
                continue
            if mime_type not in SUPPORTED_FILE_MIME_TYPES:
                continue

            drive_files.append(
                DriveFile(
                    file_id=str(candidate["id"]),
                    name=str(candidate["name"]),
                    mime_type=mime_type,
                    modified_time=parse_drive_time(str(candidate["modifiedTime"])),
                    folder_segments=folder_segments,
                )
            )

    filtered_drive_files = [
        drive_file
        for drive_file in drive_files
        if drive_file.folder_segments and drive_file.folder_segments[0] == ACTIVE_DRIVE_YEAR_FOLDER
    ]

    return sorted(filtered_drive_files, key=lambda drive_file: (drive_file.modified_time, drive_file.export_name))


def list_archive_relative_paths(archive_root: Path) -> set[str]:
    if not archive_root.exists():
        return set()
    return {
        path.relative_to(archive_root).as_posix()
        for path in archive_root.rglob("*.xlsx")
        if path.is_file()
    }


def download_drive_file(service, drive_file: DriveFile, output_path: Path) -> None:
    try:
        from googleapiclient.http import MediaIoBaseDownload
    except ImportError as exc:
        raise SystemExit(
            "Google Drive libraries are missing.\n"
            "Install them with:\n"
            "  python -m pip install -r .\\pipeline\\requirements.txt"
        ) from exc

    output_path.parent.mkdir(parents=True, exist_ok=True)

    if drive_file.mime_type == GOOGLE_SHEETS_MIME:
        request = service.files().export_media(fileId=drive_file.file_id, mimeType=XLSX_MIME)
    else:
        request = service.files().get_media(fileId=drive_file.file_id)

    with output_path.open("wb") as handle:
        downloader = MediaIoBaseDownload(handle, request)
        done = False
        consecutive_failures = 0
        max_consecutive_failures = 5

        while not done:
            try:
                record_drive_request("media_download_requests")
                _, done = downloader.next_chunk()
                consecutive_failures = 0

            except (TimeoutError, socket.timeout, ssl.SSLError) as exc:
                consecutive_failures += 1

                if consecutive_failures >= max_consecutive_failures:
                    raise

                wait_seconds = min(60, 5 * consecutive_failures)
                print(
                    f"Drive download failed during chunk for {drive_file.export_name}: "
                    f"{type(exc).__name__}. "
                    f"Retrying {consecutive_failures}/{max_consecutive_failures} "
                    f"in {wait_seconds}s..."
                )
                time.sleep(wait_seconds)

    timestamp = drive_file.modified_time.timestamp()
    os.utime(output_path, (timestamp, timestamp))


def run_subprocess(
    command: list[str], *, cwd: Path, stream_output: bool = False
) -> subprocess.CompletedProcess[str]:
    if stream_output:
        result = subprocess.run(
            command,
            cwd=str(cwd),
            text=True,
        )
        stdout = ""
        stderr = ""
    else:
        result = subprocess.run(
            command,
            cwd=str(cwd),
            text=True,
            capture_output=True,
        )
        stdout = result.stdout
        stderr = result.stderr
    if result.returncode != 0:
        raise RuntimeError(
            f"Command failed with exit code {result.returncode}: {subprocess.list2cmdline(command)}\n"
            f"stdout:\n{stdout}\n"
            f"stderr:\n{stderr}"
        )
    return subprocess.CompletedProcess(
        args=result.args,
        returncode=result.returncode,
        stdout=stdout,
        stderr=stderr,
    )


def git_command_args(repo_root: Path, *args: str) -> list[str]:
    return [
        "git",
        "-c",
        f"safe.directory={repo_root.as_posix()}",
        "-C",
        str(repo_root),
        *args,
    ]


def run_git(*args: str) -> str:
    result = run_subprocess(git_command_args(PROJECT_ROOT, *args), cwd=PROJECT_ROOT)
    return result.stdout.strip()


def current_branch() -> str:
    return run_git("rev-parse", "--abbrev-ref", "HEAD")


def tracked_changed_files(*paths: str) -> set[str]:
    commands: list[list[str]] = [
        ["diff", "--name-only", "--relative"],
        ["diff", "--cached", "--name-only", "--relative"],
    ]
    changed: set[str] = set()
    for command in commands:
        full_command = [*command]
        if paths:
            full_command.extend(["--", *paths])
        output = run_git(*full_command)
        for line in output.splitlines():
            if line.strip():
                changed.add(line.strip())
    return changed


def build_commit_message(template: str, state: dict[str, Any]) -> str:
    downloaded_files = state.get("downloaded_files")
    if isinstance(downloaded_files, list) and downloaded_files:
        workbook_count = len(downloaded_files)
        workbook_name = f"last {workbook_count} MTGO events"
        relative_path = workbook_name
    else:
        workbook_name = str(state.get("downloaded_workbook_name") or state.get("archive_workbook_name") or "workbook")
        relative_path = str(state.get("archive_relative_path") or state.get("downloaded_relative_path") or workbook_name)
    modified_date = str(state.get("downloaded_modified_date") or "")
    return template.format(
        workbook_name=workbook_name,
        relative_path=relative_path,
        modified_date=modified_date,
    )
